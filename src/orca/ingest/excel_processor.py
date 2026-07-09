"""ORCA Excel processor.

Turns an Excel workbook into clean, correctly-typed tables — one per sheet.

This file currently covers STAGES 1-3 of the 7-stage design:
  1. OPEN & INVENTORY  - open the workbook, list its sheets, read each grid.
  2. FIND THE TABLE    - locate the real header row and the data below it
                         (not a blind "row 1" - tolerates a title banner on top).
  3. READ WITH TYPES   - dates stay dates, numbers stay numbers, text stays text;
                         "NA" / blank both become a real empty (None).

Stages 4-7 (flag total rows, tag columns, split into SQL + vectors, store) come next.

The output is a `WorkbookExtract`: a plain data object the later stages read from.
"""

from __future__ import annotations

import datetime as _dt
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# Cell text that means "there is no value here" - treated the same as blank.
_EMPTY_TOKENS = {"", "na", "n/a", "n.a.", "-", "--", "none", "null"}

# How many rows from the top we scan while hunting for the header row.
_HEADER_SCAN_LIMIT = 15


class ExcelProcessingError(Exception):
    """Raised when an Excel workbook cannot be read."""


# =============================================================================
# THE SHAPE OF THE RESULT  (what stages 4-7 will read)
# =============================================================================
@dataclass
class AuxBlock:
    """A SECONDARY block found in a sheet, kept aside from the main table.

    Example: a little "Fast Calculation" summary sitting below the real data.
    We do NOT mix these into the main table's rows or its SQL numbers; we keep
    the text so it can still be searched later as context.
    """
    sheet_name: str
    start_row: int                  # 1-based Excel row where the side block starts
    end_row: int                    # 1-based Excel row where it ends
    text: str                       # the block's non-blank cells, flattened to text


@dataclass
class SheetTable:
    """One clean table extracted from one worksheet."""
    sheet_name: str
    header_row: int                 # 1-based Excel row number of the header
    columns: list[str]              # cleaned column names, left-to-right
    rows: list[dict[str, Any]]      # each row: {column_name: typed_value_or_None}
    source_rows: list[int]          # the Excel row number each data row came from
    note: str = ""                  # e.g. "empty sheet" / "no header found"
    aux_blocks: list[AuxBlock] = field(default_factory=list)  # side blocks kept aside
    aggregate_rows: list[int] = field(default_factory=list)   # indices of Total/summary rows
    column_meta: dict[str, dict] = field(default_factory=dict)  # per-column tags (stage 5)
    quality_issues: list[str] = field(default_factory=list)   # stage 6: source-data problems
                                    # found at ingestion (impossible dates, text in a
                                    # date column...). NEVER auto-corrected — reported,
                                    # so the user can fix the source file (item #11).

    @property
    def n_rows(self) -> int:
        return len(self.rows)

    @property
    def data_records(self) -> list[dict[str, Any]]:
        """The real data rows only — Total/summary rows excluded (safe to sum)."""
        agg = set(self.aggregate_rows)
        return [r for i, r in enumerate(self.rows) if i not in agg]


@dataclass
class WorkbookExtract:
    """The whole workbook after stages 1-3."""
    file_path: str
    sheets: list[SheetTable] = field(default_factory=list)


# =============================================================================
# STAGE 2 HELPERS - find the table inside a sheet
# =============================================================================
def _is_blank(value: Any) -> bool:
    """True if a cell is empty or holds a 'no value' token like 'NA' or '-'."""
    if value is None:
        return True
    if isinstance(value, str) and value.strip().lower() in _EMPTY_TOKENS:
        return True
    return False


def _find_header_row(grid: list[tuple]) -> int | None:
    """Return the 0-based index of the row that looks like column headers.

    A header row is the first row (within the top `_HEADER_SCAN_LIMIT` rows) that:
      - has at least 2 non-blank cells, AND
      - is mostly TEXT labels (headers are words, not numbers or dates).
    This skips a title banner on top (usually 1 non-blank cell) and any blank rows.
    """
    for i, row in enumerate(grid[:_HEADER_SCAN_LIMIT]):
        non_blank = [c for c in row if not _is_blank(c)]
        if len(non_blank) < 2:
            continue  # blank row or a one-cell title banner
        text_like = [c for c in non_blank if isinstance(c, str)]
        # headers are mostly words: require ~60%+ of the filled cells to be text
        if len(text_like) >= max(2, int(len(non_blank) * 0.6)):
            return i
    return None


def _column_span(header: tuple) -> tuple[int, int]:
    """First and last column index (0-based, inclusive) that the header covers."""
    filled = [idx for idx, c in enumerate(header) if not _is_blank(c)]
    return (filled[0], filled[-1])


# --- Block detection: a sheet can hold more than one table -------------------
# We cut the sheet into blocks wherever there is a fully-blank separator row,
# then keep the biggest block (+ any narrower blocks that fit inside its columns)
# as the MAIN table, and set the rest aside as side blocks (aux).

def _split_raw_blocks(grid: list[tuple]) -> list[list[tuple[int, tuple]]]:
    """Cut the grid into blocks separated by fully-blank rows.

    Each block is a list of (excel_row_number, row_values) pairs.
    """
    blocks: list[list[tuple[int, tuple]]] = []
    current: list[tuple[int, tuple]] = []
    for i, row in enumerate(grid):
        if all(_is_blank(c) for c in row):
            if current:                       # a blank row closes the current block
                blocks.append(current)
                current = []
        else:
            current.append((i + 1, row))      # i+1 = 1-based Excel row number
    if current:
        blocks.append(current)
    return blocks


def _block_span(block: list[tuple[int, tuple]]) -> tuple[int | None, int | None]:
    """Leftmost and rightmost non-blank column index (0-based) used in a block."""
    lo = hi = None
    for _row_no, row in block:
        for idx, cell in enumerate(row):
            if not _is_blank(cell):
                lo = idx if lo is None else min(lo, idx)
                hi = idx if hi is None else max(hi, idx)
    return lo, hi


def _flatten_block(block: list[tuple[int, tuple]]) -> str:
    """Turn a side block into readable text (non-blank cells only)."""
    lines = []
    for _row_no, row in block:
        cells = [str(c).strip() for c in row if not _is_blank(c)]
        if cells:
            lines.append(" | ".join(cells))
    return "\n".join(lines)


def _cell_kind(value: Any) -> str:
    """Rough type of a single cell: 'number' / 'date' / 'text'."""
    if isinstance(value, bool):          # bool is a subclass of int - treat as text
        return "text"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, _dt.datetime):
        return "date"
    return "text"


def _primary_column_types(primary: list[tuple[int, tuple]], header_idx: int,
                          lo: int, hi: int) -> dict[int, str]:
    """For the MAIN block, decide each column's dominant type from its data rows.

    Returns {column_index: 'number'|'date'|'text'|'empty'}. Used to test whether
    another block is really part of the same table.
    """
    types: dict[int, str] = {}
    for idx in range(lo, hi + 1):
        counts = {"number": 0, "date": 0, "text": 0}
        for pos, (_row_no, row) in enumerate(primary):
            if pos <= header_idx:            # skip header + any title rows above it
                continue
            cell = row[idx] if idx < len(row) else None
            if _is_blank(cell):
                continue
            counts[_cell_kind(cell)] += 1
        types[idx] = max(counts, key=counts.get) if any(counts.values()) else "empty"
    return types


def _block_matches_main(block: list[tuple[int, tuple]], col_types: dict[int, str],
                        lo: int, hi: int) -> bool:
    """True if this block's values fit the MAIN table's column types.

    A block is a DIFFERENT table (returns False) if it spills into new columns, or
    if it puts text where the main table has numbers/dates (e.g. 'Ads budget' in a
    column that is normally an amount).
    """
    b_lo, b_hi = _block_span(block)
    if b_lo is None:
        return True
    if b_lo < lo or b_hi > hi:               # uses columns the main table doesn't
        return False
    for _row_no, row in block:
        for idx in range(lo, hi + 1):
            cell = row[idx] if idx < len(row) else None
            if _is_blank(cell):
                continue
            want = col_types.get(idx, "empty")
            if want in ("number", "date") and _cell_kind(cell) != want:
                return False                 # text sitting in a numeric/date column
    return True


def _clean_column_names(header_cells: list[Any]) -> list[str]:
    """Turn raw header cells into clean, unique column names.

    Empty header cells become 'column_N'; duplicates get a '_2', '_3' suffix.
    """
    names: list[str] = []
    seen: dict[str, int] = {}
    for idx, cell in enumerate(header_cells):
        base = str(cell).strip() if not _is_blank(cell) else f"column_{idx + 1}"
        if base in seen:
            seen[base] += 1
            base = f"{base}_{seen[base]}"
        else:
            seen[base] = 1
        names.append(base)
    return names


# =============================================================================
# STAGE 3 HELPER - normalise a single cell to a clean typed value
# =============================================================================
def _clean_value(value: Any) -> Any:
    """Return a clean typed value: number / datetime / text, or None if empty.

    openpyxl already hands us native Python types (int, float, datetime, str),
    so we mostly just clean up: blanks and 'NA'-style tokens become None, and
    stray whitespace is trimmed off text.
    """
    if _is_blank(value):
        return None
    if isinstance(value, str):
        return value.strip()
    # int, float, datetime, bool, etc. pass through unchanged (types preserved)
    return value


# =============================================================================
# STAGE 4 - flag Total / summary rows (so they never double-count)
# =============================================================================
_AGG_KEYWORDS = {
    "total", "totals", "grand total", "subtotal", "sub total", "sub-total",
    "sum", "average", "avg", "mean", "overall", "summary", "ytd",
}
# We only treat a SHORT text cell as a total label - a long remark like
# "total amount still to pay" must not trip the detector.
_AGG_PREFIXES = ("total", "grand total", "subtotal", "sub-total", "average", "overall", "summary")


def _looks_aggregate(text: str) -> bool:
    """True if a short text cell reads like a 'Total' / 'Average' label."""
    t = text.strip().lower()
    if len(t) > 25:
        return False
    if t in _AGG_KEYWORDS:
        return True
    return any(t.startswith(p) for p in _AGG_PREFIXES)


def _detect_aggregate_rows(rows: list[dict[str, Any]]) -> list[int]:
    """Return the indices of rows that are Total / summary rows (any column)."""
    flagged = []
    for i, row in enumerate(rows):
        if any(isinstance(v, str) and _looks_aggregate(v) for v in row.values()):
            flagged.append(i)
    return flagged


# Detector 2 (Session 13): read the FORMULA behind each cell. A cell computed
# by =SUM(...) / =SUBTOTAL(...) over a VERTICAL range (many rows) is a total
# cell, label or no label. A row-wise formula like =SUM(F5:J5) (one row,
# several columns) is a normal data cell and must NOT trip this.
_AGG_FUNCS = re.compile(r"\b(SUM|SUBTOTAL|AVERAGE|AVERAGEA|COUNT|COUNTA|MAX|MIN)\s*\(",
                        re.IGNORECASE)
_CELL_RANGE = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)\s*:\s*\$?([A-Z]{1,3})\$?(\d+)")
_FULL_COLUMN = re.compile(r"\$?[A-Z]{1,3}\s*:\s*\$?[A-Z]{1,3}(?![A-Z0-9])")
# A range pointing at ANOTHER sheet, like Sales!B2:B715 or 'Client List'!A:A.
# A monthly Summary row that SUMs the detail sheets is a DATA row of its own
# table, not a total row — so sheet-qualified ranges must not count.
_OTHER_SHEET_REF = re.compile(
    r"(?:'[^']*'|[A-Za-z_][\w.]*)!\$?[A-Z]{1,3}(?:\$?\d+)?"
    r"(?:\s*:\s*\$?[A-Z]{1,3}(?:\$?\d+)?)?"
)


def _is_vertical_aggregate(formula: str) -> bool:
    """True if the formula aggregates DOWN a column OF ITS OWN SHEET (a total),
    not across a row and not over a different sheet."""
    if not _AGG_FUNCS.search(formula):
        return False
    own = _OTHER_SHEET_REF.sub("", formula)   # drop other-sheet ranges first
    if _FULL_COLUMN.search(own):              # =SUM(B:B) — whole own column
        return True
    for m in _CELL_RANGE.finditer(own):       # =SUM(B2:B715) — rows differ
        if m.group(2) != m.group(4):
            return True
    return False


def _formula_aggregate_rows(ws) -> set[int]:
    """1-based Excel row numbers holding at least one vertical-aggregate formula."""
    flagged: set[int] = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        for cell in row:
            if isinstance(cell, str) and cell.startswith("=") and _is_vertical_aggregate(cell):
                flagged.add(i + 1)
                break
    return flagged


# Detector 3 (Session 13, fallback for hand-typed totals with no formula):
# a mostly-empty row whose numeric cell equals its column's sum over all the
# other rows is a total row by SHAPE (same lesson as the PDF table stage).
def _detect_sum_shaped_rows(columns: list[str], rows: list[dict[str, Any]],
                            flagged: set[int]) -> list[int]:
    extra = []
    for i, row in enumerate(rows):
        if i in flagged:
            continue
        filled = [c for c in columns if row.get(c) is not None]
        if not filled or len(filled) > max(1, len(columns) // 2):
            continue                          # a real data row is mostly filled
        for col in filled:
            v = row.get(col)
            if not isinstance(v, (int, float)) or isinstance(v, bool) or v == 0:
                continue
            others = sum(
                r2.get(col) for j, r2 in enumerate(rows)
                if j != i and j not in flagged
                and isinstance(r2.get(col), (int, float))
                and not isinstance(r2.get(col), bool)
            )
            if abs(v - others) <= 0.02:       # its value IS the column's sum
                extra.append(i)
                break
    return extra


# =============================================================================
# STAGE 5 - tag each column's meaning (semantic type + aliases + real data type)
# =============================================================================
def _column_data_type(rows: list[dict[str, Any]], col: str, skip: set[int]) -> str:
    """The dominant real type of a column's values: number / date / text / empty."""
    counts = {"number": 0, "date": 0, "text": 0}
    for i, row in enumerate(rows):
        if i in skip:
            continue
        v = row.get(col)
        if v is None:
            continue
        counts[_cell_kind(v)] += 1
    return max(counts, key=counts.get) if any(counts.values()) else "empty"


def _tag_columns(columns: list[str], rows: list[dict[str, Any]],
                 aggregate_idx: list[int]) -> dict[str, dict]:
    """Give each column a semantic tag + aliases, plus its REAL data type.

    The semantic tag is guessed from the column NAME (so "profit" can later match a
    column called "Net Profit"); the data type is measured from the actual VALUES
    (more reliable than the name for deciding what is truly numeric).
    """
    skip = set(aggregate_idx)
    meta: dict[str, dict] = {}
    for col in columns:
        low = str(col).lower()
        aliases = [col]
        semantic = "text"

        if any(t in low for t in ("sales", "revenue", "income", "turnover")):
            semantic = "financial_metric"; aliases += ["sales", "revenue", "income", "turnover"]
        if any(t in low for t in ("profit", "margin", "earnings", "net")):
            semantic = "financial_metric"; aliases += ["profit", "margin", "earnings", "net profit"]
        if any(t in low for t in ("expense", "cost", "spending", "amount", "price", "deposit", "balance")):
            semantic = "financial_metric"; aliases += ["expense", "cost", "amount", "price"]
        if any(t in low for t in ("date", "month", "year", "quarter", "time", "day")):
            semantic = "temporal"; aliases += ["date", "time", "period", "month", "year"]
        if any(t in low for t in ("count", "quantity", "qty", "number", "invoices", "orders", "basket")):
            semantic = "quantity"; aliases += ["count", "quantity", "number"]
        if "%" in str(col) or "percent" in low:
            semantic = "percentage"   # checked last so "% Profit" wins over "profit"

        data_type = _column_data_type(rows, col, skip)
        meta[col] = {
            "semantic_type": semantic,
            "aliases": sorted(set(aliases)),
            "data_type": data_type,
            "is_numeric": data_type == "number",
        }
    return meta


# =============================================================================
# STAGE 6 - data-quality scan (report, NEVER auto-correct)
# =============================================================================
# Face-value rule (Walid, Session 13): the VALUE a cell shows is what we store.
# Formulas are only read to UNDERSTAND a value (e.g. "this is a vertical SUM ->
# total row"). When a face value itself is broken — a date stored in year 1900,
# text like "23 apri" sitting in a date column — we must not guess a correction:
# we STORE it as-is and REPORT it, so the user fixes the source file.
_MIN_PLAUSIBLE_YEAR = 2000          # business data before 2000 = a mistyped date


def _scan_quality(sheet_name: str, columns: list[str], rows: list[dict[str, Any]],
                  source_rows: list[int], aggregate_rows: list[int],
                  column_meta: dict[str, dict]) -> list[str]:
    """Find source-data problems in date columns. Returns plain-English notes."""
    issues: list[str] = []
    agg = set(aggregate_rows)
    date_cols = [c for c in columns if column_meta.get(c, {}).get("data_type") == "date"]
    for col in date_cols:
        for i, row in enumerate(rows):
            if i in agg:
                continue
            v = row.get(col)
            excel_row = source_rows[i]
            if isinstance(v, _dt.datetime) and v.year < _MIN_PLAUSIBLE_YEAR:
                issues.append(
                    f"{sheet_name}!row {excel_row}: '{col}' is {v.date()} — a year-"
                    f"{v.year} date is almost surely a data-entry slip (Excel kept "
                    f"the day/month but not the intended year). Fix the year in the file."
                )
            elif isinstance(v, str):
                issues.append(
                    f"{sheet_name}!row {excel_row}: '{col}' holds text {v!r} instead "
                    f"of a real date — date filters (e.g. 'March 2026') will miss "
                    f"this row until it is retyped as a date."
                )
    return issues


# =============================================================================
# STAGE 1-3 - the main entry point
# =============================================================================
def extract_workbook(excel_path: str | Path) -> WorkbookExtract:
    """Open an Excel file and return one clean, typed table per sheet (stages 1-3)."""
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    logger.info("Opening workbook: %s", path.name)
    try:
        # data_only=True  -> read the VALUES Excel last calculated (not the formulas)
        # read_only=True  -> stream large sheets efficiently (Sales has 700+ rows)
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:  # openpyxl raises various types for bad files
        raise ExcelProcessingError(f"Could not open Excel file: {e}") from e

    # Session 13: a second, formula-mode pass over the same file. Rows whose
    # cells are computed by a vertical =SUM(...)/=SUBTOTAL(...) are total rows
    # even when they carry no "Total" label (that unlabelled grand-total row in
    # the Orders/Sales sheets is exactly what the keyword detector missed).
    formula_rows: dict[str, set[int]] = {}
    try:
        wb_formulas = openpyxl.load_workbook(path, data_only=False, read_only=True)
        for sheet_name in wb_formulas.sheetnames:
            formula_rows[sheet_name] = _formula_aggregate_rows(wb_formulas[sheet_name])
        wb_formulas.close()
    except Exception as e:                     # never let the extra pass kill ingestion
        logger.warning("Formula pass failed (%s) - relying on label + shape detectors", e)

    result = WorkbookExtract(file_path=str(path))

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # STAGE 1: read the whole used grid as rows of plain values.
            grid = list(ws.iter_rows(values_only=True))
            # Some sheets carry thousands of empty trailing rows - drop them so we
            # don't waste time/memory processing blank space.
            while grid and all(_is_blank(c) for c in grid[-1]):
                grid.pop()
            result.sheets.append(
                _extract_sheet(sheet_name, grid, formula_rows.get(sheet_name, set()))
            )
    finally:
        wb.close()

    logger.info("Extracted %d sheet(s) from %s", len(result.sheets), path.name)
    return result


def _extract_sheet(sheet_name: str, grid: list[tuple],
                   formula_agg_rows: set[int] | None = None) -> SheetTable:
    """Stages 2-3 for one sheet: split into blocks, read the MAIN table, keep the
    side blocks aside. `formula_agg_rows` = 1-based Excel rows the formula pass
    flagged as vertical aggregates (totals)."""
    # Empty sheet.
    if not grid or all(all(_is_blank(c) for c in row) for row in grid):
        logger.warning("Sheet %r is empty - skipping", sheet_name)
        return SheetTable(sheet_name, 0, [], [], [], note="empty sheet")

    # STAGE 2a: cut the sheet into blocks at fully-blank separator rows.
    blocks = _split_raw_blocks(grid)

    # STAGE 2b: the MAIN table is the biggest block. We learn the main table's
    # per-column types, then keep any other block that MATCHES those types (e.g. a
    # "Total" row). A block that puts text where the main table has numbers - a
    # separate mini-table like "Fast Calculation" - is set aside as a side block.
    primary = max(blocks, key=len)
    p_lo, p_hi = _block_span(primary)
    p_header_idx = _find_header_row([row for _r, row in primary])
    if p_header_idx is None:
        p_header_idx = -1                                 # no header -> treat all as data
    col_types = _primary_column_types(primary, p_header_idx, p_lo, p_hi)

    main_rows: list[tuple[int, tuple]] = []
    aux_blocks: list[AuxBlock] = []
    for block in blocks:
        if block is primary or _block_matches_main(block, col_types, p_lo, p_hi):
            main_rows.extend(block)                       # belongs to the main table
        else:
            aux_blocks.append(AuxBlock(
                sheet_name=sheet_name,
                start_row=block[0][0],
                end_row=block[-1][0],
                text=_flatten_block(block),
            ))
    main_rows.sort(key=lambda rc: rc[0])                  # keep Excel row order

    # STAGE 2c: find the header row within the assembled main table.
    values_only = [row for _row_no, row in main_rows]
    header_idx = _find_header_row(values_only)
    if header_idx is None:
        logger.warning("Sheet %r: no clear header row found", sheet_name)
        return SheetTable(sheet_name, 0, [], [], [], note="no header found",
                          aux_blocks=aux_blocks)

    header_excel_row, header_row = main_rows[header_idx]
    first_col, last_col = _column_span(header_row)
    columns = _clean_column_names(list(header_row[first_col:last_col + 1]))

    # STAGE 3: read every data row BELOW the header, cleaned + typed.
    rows: list[dict[str, Any]] = []
    source_rows: list[int] = []
    for excel_row, raw_row in main_rows[header_idx + 1:]:
        cells = list(raw_row[first_col:last_col + 1])
        cells += [None] * (len(columns) - len(cells))     # pad short rows
        if all(_is_blank(c) for c in cells):              # skip a stray blank row
            continue
        record = {col: _clean_value(cells[i]) for i, col in enumerate(columns)}
        rows.append(record)
        source_rows.append(excel_row)                     # for later citations

    # STAGE 4: flag Total / summary rows so they never double-count in sums.
    # Three detectors, union of all (Session 13):
    #   1. label   — a short "Total"/"Average" text cell
    #   2. formula — the cell is computed by a vertical =SUM(...) (no label needed)
    #   3. shape   — mostly-empty row whose number equals its column's sum
    flagged = set(_detect_aggregate_rows(rows))
    if formula_agg_rows:
        flagged |= {i for i, excel_row in enumerate(source_rows)
                    if excel_row in formula_agg_rows}
    flagged |= set(_detect_sum_shaped_rows(columns, rows, flagged))
    aggregate_rows = sorted(flagged)
    # STAGE 5: tag each column's meaning + real data type.
    column_meta = _tag_columns(columns, rows, aggregate_rows)
    # STAGE 6: scan for source-data problems (stored as-is, reported to the user).
    quality_issues = _scan_quality(sheet_name, columns, rows, source_rows,
                                   aggregate_rows, column_meta)

    logger.info(
        "Sheet %r: header at row %d, %d columns, %d data rows (%d total-rows), %d side block(s)",
        sheet_name, header_excel_row, len(columns), len(rows),
        len(aggregate_rows), len(aux_blocks),
    )
    return SheetTable(
        sheet_name=sheet_name,
        header_row=header_excel_row,
        columns=columns,
        rows=rows,
        source_rows=source_rows,
        aux_blocks=aux_blocks,
        aggregate_rows=aggregate_rows,
        column_meta=column_meta,
        quality_issues=quality_issues,
    )
